import tkinter
from tkinter import ttk
import re
import os
import sqlite3
import time
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import ast
import re
import ast



# L:500,B:430,H:855  ==>  {"L":"500","B":"430","H":"855"}
def split_size_table(stroka):
    try:
        if stroka == "":
            return {"size": "?"}
        Return_Dictionary = {}
        K_V_list = stroka.split(',')
        for each in K_V_list:
            Return_Dictionary[each.split(":")[0]] = each.split(":")[1]
        return Return_Dictionary
    except Exception as e:
        return {"-":"-"}



def splitter(stroka = str):
    # stroka = List_cleaner(stroka)
    # print(stroka)
    X = []
    Y = []
    # если введенные точки явно откуда-то скопированы, то там не иначе как будет переход на другую строку
    if '\n' in stroka:
        spisok = stroka.split('\n')
        # если разделители такие, то запятая будет делить целое и не целое
        if '\t' in stroka or ';' in stroka:
            for each in spisok:
                # print(each)
                if ";" in each:
                    X.append(each.split(";")[0])
                    Y.append(each.split(";")[1])
                elif "\t" in each:
                    X.append(each.split("\t")[0])
                    Y.append(each.split("\t")[1])
            if "," in X[0]:
                X = [float(re.sub(r',', '.', element)) for element in X]
                Y = [float(re.sub(r',', '.', element)) for element in Y]
        # Если есть одновременно и запятая и точка. То запятая делит X, Y
        elif ',' in stroka and '.' in stroka:
            for each in spisok:
                try:
                    X.append(float(each.split(",")[0]))
                    Y.append(float(each.split(",")[1]))
                except:
                    pass
        elif ',' in stroka and not '.' in stroka:
            for each in spisok:
                try:
                    # print(each)
                    # print(each.split(","))
                    # X.append(float(each.split(",")[0]))
                    Y.append(float(re.sub(r',', ".", each, flags=re.DOTALL)))
                except:
                    pass
    elif ',' in stroka  and '\n' not in stroka:
        if "!" in stroka:
            xy = stroka.split("!")
            xspisok = xy[0].split(',')
            for each in xspisok:
                try:
                    X.append(float(each))
                except:
                    pass
            yspisok = xy[1].split(',')
            for each in yspisok:
                try:
                    Y.append(float(each))
                except:
                    pass
            pass
        else:
            spisok = stroka.split(',')
            for each in spisok:

                try:
                    Y.append(float(each))
                except:
                    pass
    # Если была введена жесть, то она выкидывает те пары значений, которые не цифры
    if len(X)>=1 and type(X[0]) == str:
        # while str in X:
        indexator = 0
        while indexator<len(X) or indexator<len(Y):
            try:
                X[indexator] = float(re.sub(r',', '.', X[indexator], flags=re.DOTALL))
                Y[indexator] = float(re.sub(r',', '.', Y[indexator], flags=re.DOTALL))
                indexator+=1
            except:
                X.pop(indexator)
                Y.pop(indexator)
    try:
        promezhutok = ast.literal_eval(stroka)
        if type(promezhutok) != tuple:
            if len(promezhutok) == 2:
                X = promezhutok[0]
                Y = promezhutok[1]
            else:
                Y = promezhutok
        return X, Y
    except:
        pass
    return X, Y


class DB_manager:
    def RunQueryNoReturn(self, SQLQuery, Values=None, maxRetry=1):
        not_done_flag = False
        maxRetry = maxRetry
        con = sqlite3.connect(self.DBName)
        Cursor = con.cursor()
        if Values != None:
            Cursor.execute(SQLQuery, Values)
        else:
            Cursor.execute(SQLQuery)
        con.commit()
        Cursor.close()
        con.close()

    def RunQuery(self, SQLQuery, maxRetry=1):
        not_done_flag = False
        maxRetry = maxRetry
        con = sqlite3.connect(self.DBName)
        Cursor = con.cursor()
        Cursor.execute(SQLQuery)
        results = Cursor.fetchall()
        con.commit()
        Cursor.close()
        con.close()
        return results




    def tableWithSingleValueOperation(self, TableName,Operation, newValue = ""):
    # create, get, set, update
        if Operation == "create":
            AdminKeyTable = """CREATE TABLE IF NOT EXISTS {} (id_ INTEGER, value TEXT)""".format(TableName)
            self.RunQueryNoReturn(AdminKeyTable)
        if Operation == "set":
            try:
                DeleteQuery = f"DELETE FROM {TableName} WHERE id_ ='{0}' "
                self.RunQueryNoReturn( DeleteQuery)
            finally:
                SQLquery = f"INSERT INTO {TableName}  VALUES(?,?)"
                self.RunQueryNoReturn(SQLquery, Values=(0,str(newValue)))
        if Operation == "get":
            try:
                table = self.RunQuery("SELECT * FROM {}".format(TableName))
                return table[0][1]
            except:
                self.tableWithSingleValueOperation(TableName,"set", newValue = "")
                table = self.RunQuery("SELECT * FROM {}".format(TableName))
                return table[0][1]


        pass
    def giveColumnDescription(self, TableName, what = "Names"):
        parametr = 1
        if what == "Names":
            parametr =1
        elif what =="Types":
            parametr = 2
        String = "PRAGMA table_info({});".format(TableName)
        columnNames = []
        for columnDescription in self.RunQuery(String):
            columnNames.append(columnDescription[parametr])
        return columnNames

    def ReadXLS(self, fileName, max_columns=None, max_rows=None):
        filename = '%s.xlsx' % (fileName)
        wb = load_workbook(filename=filename, read_only=False)
        ws = wb.active
        XLSTable = []
        if max_columns == None:
            max_columns = ws.max_column
        if max_rows == None:
            max_rows = ws.max_row
        for row in ws.iter_rows(min_row=1, max_col=max_columns, max_row=max_rows):
            rowList = []
            for cell in row:
                rowList.append(cell.value)
            XLSTable.append(rowList)
        return XLSTable

    def writeXLS(self, fileName):
        titles = self.giveColumnDescription(fileName, what = "Names")
        data = self.RunQuery("SELECT * FROM {}".format(fileName))
        filename = '%s.xlsx' % (fileName)
        wb = Workbook()
        ws = wb.active
        ws.append(titles)
        wb.save(filename)
        wb.close()

        wb = load_workbook(filename=filename, read_only=False)
        ws = wb.active
        for each in data:
            ws.append(each)
        wb.save(filename)
        wb.close()
        return filename

    def CreateTableWithThisXLS(self, FileAndTableName="Pumpsxlsx"):

        XLTable = self.ReadXLS(FileAndTableName)
        for row in XLTable:
            # если строка первая, то там названия колонок, так что мы создаем таблицу и создаем эти колонки какие есть
            if XLTable.index(row) == 0:
                ColumnNamesString = ''
                for each in range(len(row)):

                    typeOfTHisColumn = " TEXT "
                    try:
                        itsInt = str(XLTable[1][each])
                        if str(type((itsInt))) == "<class 'str'>":
                            typeOfTHisColumn = " TEXT "
                    except Exception as e:
                        print("Ошибка DBFunctions ", e)
                        print("возмоожно  excel пустая или есть пустые подозрительные места")
                    ColumnNamesString = ColumnNamesString + " " + row[each] + typeOfTHisColumn
                    if row.index(row[each]) < len(row) - 1:
                        ColumnNamesString = ColumnNamesString + ","
                Table = """CREATE TABLE IF NOT EXISTS {} (
                                {}
                                )""".format(FileAndTableName, ColumnNamesString)
                # print(Table)
                self.RunQueryNoReturn(Table)
                continue

            # дальше мы построчно добавляем строки
            amount = ""
            for each in range(len(row)):
                amount = amount + "?"
                if row.index(row[each]) < len(row) - 1:
                    amount = amount + ","

            # print(amount)
            SQLquery = f"INSERT INTO %s VALUES(%s)" % (FileAndTableName, amount)

            Values = []

            columnTypes = self.giveColumnDescription(FileAndTableName, what = "Types")
            # print(columnTypes)
            for each in range(len(row)):
                if columnTypes[each] == 'INTEGER':
                    Values.append(int(row[each]))
                elif columnTypes[each] == 'TEXT':
                    Values.append(str(row[each]))
            print(SQLquery)
            print(Values)
            self.RunQueryNoReturn(SQLquery, Values=Values)

    def setDefaultColumnsList(self,tableName):
        Table = f"""CREATE TABLE IF NOT EXISTS {tableName}(
                Column_Name TEXT ,
                Column_Type TEXT,
                Column_Description TEXT
                )"""
        self.RunQueryNoReturn(Table)
        self.RunQueryNoReturn(f"INSERT INTO {tableName} VALUES(?,?,?)", Values=("Model_Name", "Данные о продукте", "Название модели"))


    def setDefaultBaseTable(self,tableName, basicString = True):
        Table = f"""CREATE TABLE IF NOT EXISTS {tableName}(
                Model_Name TEXT ,
                max_pressure TEXT ,
                liquid_pump_temp TEXT ,
                max_env_temp TEXT ,

                engine_efficiency_class TEXT ,
                Net_connection TEXT ,
                rotation_frequency_nominal TEXT ,
                Nominal_power_P2 TEXT ,
                Nominal_currency TEXT ,
                Moisture_protection TEXT ,
                Isolation_Class TEXT ,

                Sucking_pipe_branch TEXT ,
                Pushing_pipe_branch TEXT ,
                building_length TEXT ,

                frame TEXT ,
                working_wheel TEXT ,
                Shaft TEXT ,
                Weight TEXT ,

                lift_H_m TEXT ,
                Cavitation_NPSH_m TEXT ,
                motor_power_P2_kW TEXT ,
                efficiency TEXT ,
                Scale_Table TEXT ,
                pump_wiring_connection_figure TEXT ,
                pump_figure TEXT
                )"""
        self.RunQueryNoReturn(Table)
        if basicString == True:
            self.RunQueryNoReturn(f"INSERT INTO {tableName} VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                             Values=('model Test', ' 16 бар', ' -15C-110C', ' 40С', ' IE2', ' 3~380V/50 Hz', ' 2900 1/min', ' P2 15 кВт', ' 27 А', ' IP54', ' F', ' DN100', ' DN100', ' 480 мм', ' GG20', ' GG20', ' 1.4021', ' 335 ', '[[0.0, 11.77, 23.54, 35.31], [22.96, 22.92, 22.93, 22.92]]', '[[0.0, 11.77, 23.54, 35.31], [1.14, 1.17, 1.2, 1.24]]', '[[0.0, 11.77, 23.54, 35.31], [3.57, 3.84, 4.21, 4.67]]', '[[5.884298443, 11.76859689, 17.65289533, 23.53719377], [18.68826521, 19.21745255, 19.86845215, 20.71937448]]', "{'l': '34'}", 'connection_figures/test_connection.png', 'pump_figures/Test_pump_image.png'))




    # получает два имени таблицы - 1 таблица содержит 2 колонки в первой имена колонок во второй таблице и во второй колонке тип имени.
    # получает новый список колонок . 1 таблица изменяется в соответствии с списком и приводит колонки второй таблицы в соответствии с первой
    def editColumnsAccordingTo(self,TableName1, TableName2, NewColsList):
        if type(NewColsList) == str:
            try:
                NewColsList = ast.literal_eval(NewColsList)
            except Exception as e:
                print(e)
                pass
        def GiveColNames():
            Col1Names = self.RunQuery(f"SELECT Column_Name FROM {TableName1}")
            Col1Names = [each[0] for each in Col1Names]
            Col2Names = self.giveColumnDescription(TableName2, what="Names")
            return Col1Names, Col2Names
        Col1Names, Col2Names = GiveColNames()
        print("Col1Names",Col1Names)
        print("Col2Names",Col2Names)
        print("NewColsList",NewColsList)

        #операция приводит таблицу 1 (описания таблицы 2) в соответствии с новым списком
        #добавляем новое
        for newColumnDescription in NewColsList:
            if newColumnDescription[0] not in Col1Names:
                self.RunQueryNoReturn(f"INSERT INTO {TableName1} VALUES(?,?,?)", Values=(newColumnDescription[0],newColumnDescription[1],newColumnDescription[2]))
        # Удаляем чего нет
        NewColumns = [each[0] for each in NewColsList]
        print("NewColumns",NewColumns)#должен быть список
        for ExistingColumn in Col1Names:
            if ExistingColumn not in NewColumns:
                DeleteQuery = f"DELETE FROM {TableName1} WHERE Column_Name ='{ExistingColumn}' "
                # print(DeleteQuery)
                self.RunQueryNoReturn(DeleteQuery)

        #операция приводит колонки второй таблицы в соответствии записями в первой таблице
        #добавляем новое
        Col1Names, Col2Names = GiveColNames()
        for columnName in Col1Names:
            if columnName not in Col2Names:
                self.RunQueryNoReturn(f"ALTER TABLE {TableName2} ADD {columnName} TEXT")
        # Удаляем чего нет
        for ExistingColumn in Col2Names:
            if ExistingColumn not in Col1Names:
                quary = f"ALTER TABLE {TableName2} DROP COLUMN {ExistingColumn};"
                self.RunQueryNoReturn(quary)
                # sqlite до 3.35.0 не умеет удалять колонки, поэтому надо создать копию таблицы, удалить оригинальную, создать заново и поколоночно добавить из копии все столбцы, кроме удаляемых
                pass

    def __init__(self, DBName):
        self.DBName = DBName



# в бд в итоге три таблицы.
# 1 - таблица в одну строку и хранит ключ админа
# 2 - таблица со списком колонок третьей и к каждой колонке соответствует её тип
# 3 - непосредственно данные насосов.





class fieldAndLabel:
    myName = "UnnamedfieldAndLabel"
    label = ""
    raw = 0
    column = 0
    entryDefault = ''
    entry = None#tkinter.Entry()

    def set(self, value, alsoDefault = False):
        self.entry.delete(0, len(self.entry.get()))
        self.entry.insert(0, str(value))
        if alsoDefault == True:
            self.entryDefault = str(value)
    def content(self):
        try:
            return self.entry.get()
        except:
            print("объект "+self.myName+ " не может взять данные из поля")
            return self.myName,None
    def clear(self):
        try:
            self.entry.delete(0, len(self.entry.get()))
            self.entry.insert(0, self.entryDefault)
        except Exception as e:
            print("не получается очистить Entry в  " + self.myName)
        pass

    def __init__(self, frame, myName, Label, raw, column, entryDefault="" ):
        self.myName = myName
        self.label = Label
        self.raw = raw
        self.column = column
        self.entryDefault = entryDefault

        tkinter.Label(frame, text=self.label).grid(row=self.raw, column=self.column)
        self.entry = tkinter.Entry(frame, width=40)
        self.entry.grid(row=self.raw+1, column=self.column, columnspan=1)
        self.entry.insert(0, self.entryDefault)



class adminWindow:
    columnDescriptions = []
    EntriesList = {}
    db = None

    def fillContent(self,frame):
        def CountXY(amount, Xequals=1, Yequals=10):
            x = 0
            y = 0
            for Ys in range(1, 100):
                for Xs in range(1, 100):
                    if Ys * Yequals * Xs * Xequals >= amount:
                        y = Ys * Yequals
                        x = Xs * Xequals
                        return x, y
            return x, y
        def count(xNow,yNow):
            yNow+=1
            if yNow>X:
                yNow=0
                xNow+=1
            return xNow, yNow
        xNow,yNow = 0,0

        additionalPlacesToDB = 20
        X, Y = CountXY(len(self.columnDescriptions)+additionalPlacesToDB, Xequals=1, Yequals=4)

        flow_rate_Q_m3_h = fieldAndLabel(frame, "flow_rate_Q_m3_h", "Расход - Q(m3 / h) (исп. если не введен для H/NPSH/P2/КПД)", yNow * 2, xNow, entryDefault="")
        self.EntriesList["flow_rate_Q_m3_h"] = flow_rate_Q_m3_h
        xNow, yNow = count(xNow, yNow)

        for columnIndex in range(len(self.columnDescriptions)):
            entryObj = fieldAndLabel(frame, self.columnDescriptions[columnIndex][0], self.columnDescriptions[columnIndex][2], yNow*2, xNow, entryDefault="")
            self.EntriesList[self.columnDescriptions[columnIndex][0]] = entryObj
            xNow, yNow = count(xNow, yNow)

        disAssemble_line = fieldAndLabel(frame, "disAssemble_line", "разбить введенные данные в поля автоматически", yNow * 2, xNow, entryDefault="")
        self.EntriesList["disAssemble_line"] = disAssemble_line
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text="разобрать по формам", command=lambda: self.disAssemble_line(text=self.EntriesList["disAssemble_line"].content())).grid(row=yNow * 2, column=xNow, columnspan=1, rowspan=2)
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text=" Предпросмотр PDF ",command=lambda: self.ShowPDF()).grid(row=yNow * 2, column=xNow, columnspan=1, rowspan = 2)
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text=" Добавить запись в базу",command=lambda: self.add_to_DB()).grid(row=yNow * 2, column=xNow, columnspan=1, rowspan = 2)
        xNow, yNow = count(xNow, yNow)
        AdminKeyEntry = fieldAndLabel(frame,"AdminKeyEntry", "Ключ администратора", yNow * 2, xNow, entryDefault="")
        self.EntriesList["AdminKeyEntry"] = AdminKeyEntry
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text=" Сменить ключ администратора",command=lambda: self.updateAdminKey()).grid(row=yNow * 2, column=xNow, columnspan=1, rowspan = 2)
        xNow, yNow = count(xNow, yNow)
        Unload_from_DB_to_XLSX_entry = fieldAndLabel(frame,"Unload_from_DB_to_XLSX_entry", "Вывести таблицу из базы данных в excel", yNow * 2, xNow, entryDefault="PumpsColums")
        self.EntriesList["Unload_from_DB_to_XLSX_entry"] = Unload_from_DB_to_XLSX_entry
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text=" DB => XL",command=lambda: self.Unload_from_DB_to_XLSX()).grid(row=yNow * 2, column=xNow, columnspan=1, rowspan = 2)
        xNow, yNow = count(xNow, yNow)
        Upload_from_XLSX_to_DB_entry = fieldAndLabel(frame,"Upload_from_XLSX_to_DB_entry", "Занести таблицу в БД", yNow * 2, xNow, entryDefault="PumpsColums")
        self.EntriesList["Upload_from_XLSX_to_DB_entry"] = Upload_from_XLSX_to_DB_entry
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text=" XL => DB",command=lambda: self.Upload_from_XLSX_to_DB()).grid(row=yNow * 2, column=xNow, columnspan=1, rowspan = 2)
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text="Привести колонки таблицы насосов\nв соответствии с таблицей описания колонок", command=lambda: self.editColumnsAccordingTo()).grid(row=yNow * 2,
                                                                                                    column=xNow,
                                                                                                    columnspan=1,
                                                                                                    rowspan=2)

        try:
            self.EntriesList["pump_figure"].set('pump_figures/Test_pump_image.png', alsoDefault = True)
        except:pass
        try:
            self.EntriesList["pump_wiring_connection_figure"].set('connection_figures/test_connection.png', alsoDefault = True)
        except:pass
        pass


    #Инструменты работы внутри окна
    def Clear_all_entry(self):
        for k,v in self.EntriesList.items():
            self.EntriesList[k].clear()

    def CollectAllForms(self):
        dict_to_row = {}
        for k,v in self.EntriesList.items():
            dict_to_row[k] = self.EntriesList[k].content()
        # lift_H_m
        lift_H_mX = splitter(dict_to_row['lift_H_m'])[0]
        lift_H_mY = splitter(dict_to_row['lift_H_m'])[1]
        if lift_H_mX == []:
            lift_H_mX = splitter(dict_to_row['flow_rate_Q_m3_h'])[1]
        dict_to_row['lift_H_m'] =str([lift_H_mX,lift_H_mY])


        # Cavitation_NPSH_m
        Cavitation_NPSH_mX = splitter(dict_to_row['Cavitation_NPSH_m'])[0]
        Cavitation_NPSH_mY = splitter(dict_to_row['Cavitation_NPSH_m'])[1]
        if Cavitation_NPSH_mX == []:
            Cavitation_NPSH_mX = splitter(dict_to_row['flow_rate_Q_m3_h'])[1]
        dict_to_row['Cavitation_NPSH_m'] =str([Cavitation_NPSH_mX,Cavitation_NPSH_mY])

        motor_power_P2_kWX = splitter(dict_to_row['motor_power_P2_kW'])[0]
        motor_power_P2_kWY = splitter(dict_to_row['motor_power_P2_kW'])[1]
        if motor_power_P2_kWX == []:
            motor_power_P2_kWX = splitter(dict_to_row['flow_rate_Q_m3_h'])[1]
        dict_to_row['motor_power_P2_kW'] =str([motor_power_P2_kWX,motor_power_P2_kWY])

        # efficiency
        efficiencyX = splitter(dict_to_row['efficiency'])[0]
        efficiencyY = splitter(dict_to_row['efficiency'])[1]
        if efficiencyX == []:
            efficiencyX = splitter(dict_to_row['flow_rate_Q_m3_h'])[1]
        dict_to_row['efficiency'] =str([efficiencyX,efficiencyY])

        dict_to_row['Scale_Table'] = str(split_size_table(dict_to_row['Scale_Table']))

        print(dict_to_row['lift_H_m'])
        print(dict_to_row['Cavitation_NPSH_m'])
        print(dict_to_row['motor_power_P2_kW'])
        print(dict_to_row['efficiency'])
        print(dict_to_row['Scale_Table'])
        return dict_to_row
        pass

    def disAssemble_line(self, text="42"):
        variants = {"max_pressure":r"(?<=давление).*?(?=\n)",
                    "liquid_pump_temp":r"(?<=жидкости).*?(?=\n)",
                    "max_env_temp":r"(?<=Среды).*?(?=\n)",
                    "engine_efficiency_class":r"(?<=двигателя).*?(?=\n)",
                    "Net_connection":r"(?<=сети).*?(?=\n)",
                    'rotation_frequency_nominal':r"(?<=вращения).*?(?=\n)",
                    "Nominal_power_P2":r"(?<=мощность).*?(?=\n)",
                    "Nominal_currency":r"(?<=ток).*?(?=\n)",
                    "Moisture_protection":r"(?<=защиты).*?(?=\n)",
                    "Isolation_Class":r"(?<=изоляции).*?(?=\n)",
                    "Sucking_pipe_branch":r"(?<=всасывания).*?(?=\n)",
                    "Pushing_pipe_branch":r"(?<=напорной стороне).*?(?=\n)",
                    "building_length":r"(?<=длина).*?(?=\n)",
                    "frame":r"(?<=Корпус).*?(?=\n)",
                    "working_wheel":r"(?<=колесо).*?(?=\n)",
                    "Shaft":r"(?<=Вал).*?(?=\n)",
                    "Weight":r"(?<=Вес).*?(?=кг)"
                    }
        for k,v in variants.items():
            try:
                self.EntriesList[k].set(re.search(variants[k], text)[0])
            except:
                pass


    def ShowPDF(self):
        The_Line_dict = self.CollectAllForms()
        print(The_Line_dict)

    def add_to_DB(self):
        # собираем контент с полей
        The_Line_dict =  self.CollectAllForms()
        Columns_ordered = self.db.giveColumnDescription("Pumps", what = "Names")

        NewLineOrdered = []
        for each in Columns_ordered:
            NewLineOrdered.append(The_Line_dict[each])

        amount = ""
        for each in range(len(NewLineOrdered)):
            amount = amount + "?"
            if NewLineOrdered.index(NewLineOrdered[each]) < len(NewLineOrdered) - 1:
                amount = amount + ","

        SQLquery = f"INSERT INTO Pumps VALUES({amount})"
        print(SQLquery)
        self.db.RunQueryNoReturn(SQLquery, Values = NewLineOrdered)
        self.Clear_all_entry()

    def updateAdminKey(self):
        num = self.EntriesList["AdminKeyEntry"].content()
        self.db.tableWithSingleValueOperation("AdminKey","set", newValue = "num")
        self.EntriesList["AdminKeyEntry"].clear()
        print(num)

    def Upload_from_XLSX_to_DB(self):
        try:
            FileName = self.EntriesList["Upload_from_XLSX_to_DB_entry"].content()
            self.EntriesList["Upload_from_XLSX_to_DB_entry"].clear()
            try:
                self.db.RunQueryNoReturn(f"DROP TABLE {FileName};")
            except:
                pass
            self.db.CreateTableWithThisXLS(FileAndTableName=FileName)
            pass
        except Exception as e:
            print(e)

    def Unload_from_DB_to_XLSX(self):
        try:
            FileName = self.EntriesList["Unload_from_DB_to_XLSX_entry"].content()
            self.EntriesList["Unload_from_DB_to_XLSX_entry"].clear()
            FileName = self.db.writeXLS(FileName)
            try:
                os.startfile(FileName)
            except:
                os.system("xdg-open \"%s\"" % FileName)
            pass
        except Exception as e:
            print(e)
    def editColumnsAccordingTo(self):
        ndb = self.db.RunQuery("SELECT * FROM {}".format("PumpsColums"))
        self.db.editColumnsAccordingTo("PumpsColums", "Pumps", ndb)

    def __init__(self, dbObject):
        self.db = dbObject
        try:
            self.columnDescriptions = self.db.RunQuery("SELECT * FROM {}".format("PumpsColums"))
        except:
            self.db.setDefaultColumnsList("PumpsColums")
            self.db.setDefaultBaseTable("Pumps")
            self.columnDescriptions = self.db.RunQuery("SELECT * FROM {}".format("PumpsColums"))
        root2 = tkinter.Toplevel()
        frame = ttk.Frame(root2, padding=0)
        frame.grid()
        self.fillContent(frame)



class mainWindow:
    db = DB_manager("DB.sql")
    FormsList = []

    def OpenCalculator(self):
        for each in self.FormsList:
            print(each.content())
        pass

    def fillContent(self, frame):
        Flow = fieldAndLabel(frame, "Flow" , "Производительность(Q)", 0, 0, entryDefault = "100")
        pressure = fieldAndLabel(frame, "pressure" , "Напор(H)", 2, 0, entryDefault = "19")
        offsetX = fieldAndLabel(frame, "offsetX", "offsetX", 4, 0, entryDefault="10")
        offsetY = fieldAndLabel(frame, "offsetY", "offsetY", 6, 0, entryDefault="10")
        key = fieldAndLabel(frame, "key", "Ключ администратора", 8, 0)

        self.FormsList.append(Flow)
        self.FormsList.append(pressure)
        self.FormsList.append(offsetX)
        self.FormsList.append(offsetY)
        self.FormsList.append(key)
        tkinter.Button(frame, text="\n\n провести подбор \n\n", command =lambda: self.OpenCalculator()).grid(row=0, column=1, rowspan = 4)
        tkinter.Button(frame, text="Окно администратора", command =lambda: self.OpenAdminWindow()).grid(row=8, column=1, columnspan=1)




    def OpenAdminWindow(self):
        p = adminWindow(self.db)

    def __init__(self):
        root = tkinter.Tk()
        window1_frame = ttk.Frame(root, padding=0)
        window1_frame.grid()
        self.fillContent(window1_frame)
        root.mainloop()



if __name__ == '__main__':
    m = mainWindow()
    pass
