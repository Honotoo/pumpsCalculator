import sqlite3


import time
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import ast


class DB_manager:
    def RunQueryNoReturn(self, SQLQuery, Values=None, maxRetry=1):
        not_done_flag = False
        maxRetry = maxRetry
        con = sqlite3.connect(self.DBName)
        Cursor = con.cursor()
        if Values != None:
            Cursor.execute(SQLQuery, Values)
        else:
            Cursor.execute(SQLQuery)
        con.commit()
        Cursor.close()
        con.close()

    def RunQuery(self, SQLQuery, maxRetry=1):
        not_done_flag = False
        maxRetry = maxRetry
        con = sqlite3.connect(self.DBName)
        Cursor = con.cursor()
        Cursor.execute(SQLQuery)
        results = Cursor.fetchall()
        con.commit()
        Cursor.close()
        con.close()
        return results




    def tableWithSingleValueOperation(self, TableName,Operation, newValue = ""):
    # create, get, set, update
        if Operation == "create":
            AdminKeyTable = """CREATE TABLE IF NOT EXISTS {} (id_ INTEGER, value TEXT)""".format(TableName)
            self.RunQueryNoReturn(AdminKeyTable)
        if Operation == "set":
            try:
                DeleteQuery = f"DELETE FROM {TableName} WHERE id_ ='{0}' "
                self.RunQueryNoReturn( DeleteQuery)
            finally:
                SQLquery = f"INSERT INTO {TableName}  VALUES(?,?)"
                self.RunQueryNoReturn(SQLquery, Values=(0,str(newValue)))
        if Operation == "get":
            try:
                table = self.RunQuery("SELECT * FROM {}".format(TableName))
                return table[0][1]
            except:
                self.tableWithSingleValueOperation(TableName,"set", newValue = "")
                table = self.RunQuery("SELECT * FROM {}".format(TableName))
                return table[0][1]


        pass
    def giveColumnDescription(self, TableName, what = "Names"):
        parametr = 1
        if what == "Names":
            parametr =1
        elif what =="Types":
            parametr = 2
        String = "PRAGMA table_info({});".format(TableName)
        columnNames = []
        for columnDescription in self.RunQuery(String):
            columnNames.append(columnDescription[parametr])
        return columnNames

    def ReadXLS(self, fileName, max_columns=None, max_rows=None):
        filename = '%s.xlsx' % (fileName)
        wb = load_workbook(filename=filename, read_only=False)
        ws = wb.active
        XLSTable = []
        if max_columns == None:
            max_columns = ws.max_column
        if max_rows == None:
            max_rows = ws.max_row
        for row in ws.iter_rows(min_row=1, max_col=max_columns, max_row=max_rows):
            rowList = []
            for cell in row:
                rowList.append(cell.value)
            XLSTable.append(rowList)
        return XLSTable

    def writeXLS(self, fileName):
        titles = self.giveColumnDescription(fileName, what = "Names")
        data = self.RunQuery("SELECT * FROM {}".format(fileName))
        filename = '%s.xlsx' % (fileName)
        wb = Workbook()
        ws = wb.active
        ws.append(titles)
        wb.save(filename)
        wb.close()

        wb = load_workbook(filename=filename, read_only=False)
        ws = wb.active
        for each in data:
            ws.append(each)
        wb.save(filename)
        wb.close()
        return filename

    def CreateTableWithThisXLS(self, FileAndTableName="Pumpsxlsx"):

        XLTable = self.ReadXLS(FileAndTableName)
        for row in XLTable:
            # если строка первая, то там названия колонок, так что мы создаем таблицу и создаем эти колонки какие есть
            if XLTable.index(row) == 0:
                ColumnNamesString = ''
                for each in range(len(row)):

                    typeOfTHisColumn = " TEXT "
                    try:
                        itsInt = str(XLTable[1][each])
                        if str(type((itsInt))) == "<class 'str'>":
                            typeOfTHisColumn = " TEXT "
                    except Exception as e:
                        print("Ошибка DBFunctions ", e)
                        print("возмоожно  excel пустая или есть пустые подозрительные места")
                    ColumnNamesString = ColumnNamesString + " " + row[each] + typeOfTHisColumn
                    if row.index(row[each]) < len(row) - 1:
                        ColumnNamesString = ColumnNamesString + ","
                Table = """CREATE TABLE IF NOT EXISTS {} (
                                {}
                                )""".format(FileAndTableName, ColumnNamesString)
                # print(Table)
                self.RunQueryNoReturn(Table)
                continue

            # дальше мы построчно добавляем строки
            amount = ""
            for each in range(len(row)):
                amount = amount + "?"
                if row.index(row[each]) < len(row) - 1:
                    amount = amount + ","

            # print(amount)
            SQLquery = f"INSERT INTO %s VALUES(%s)" % (FileAndTableName, amount)

            Values = []

            columnTypes = self.giveColumnDescription(FileAndTableName, what = "Types")
            # print(columnTypes)
            for each in range(len(row)):
                if columnTypes[each] == 'INTEGER':
                    Values.append(int(row[each]))
                elif columnTypes[each] == 'TEXT':
                    Values.append(str(row[each]))
            print(SQLquery)
            print(Values)
            self.RunQueryNoReturn(SQLquery, Values=Values)

    def setDefaultColumnsList(self,tableName):
        Table = f"""CREATE TABLE IF NOT EXISTS {tableName}(
                Column_Name TEXT ,
                Column_Type TEXT,
                Column_Description TEXT
                )"""
        self.RunQueryNoReturn(Table)
        self.RunQueryNoReturn(f"INSERT INTO {tableName} VALUES(?,?,?)", Values=("Model_Name", "Данные о продукте", "Название модели"))


    def setDefaultBaseTable(self,tableName, basicString = True):
        Table = f"""CREATE TABLE IF NOT EXISTS {tableName}(
                Model_Name TEXT ,
                max_pressure TEXT ,
                liquid_pump_temp TEXT ,
                max_env_temp TEXT ,

                engine_efficiency_class TEXT ,
                Net_connection TEXT ,
                rotation_frequency_nominal TEXT ,
                Nominal_power_P2 TEXT ,
                Nominal_currency TEXT ,
                Moisture_protection TEXT ,
                Isolation_Class TEXT ,

                Sucking_pipe_branch TEXT ,
                Pushing_pipe_branch TEXT ,
                building_length TEXT ,

                frame TEXT ,
                working_wheel TEXT ,
                Shaft TEXT ,
                Weight TEXT ,

                lift_H_m TEXT ,
                Cavitation_NPSH_m TEXT ,
                motor_power_P2_kW TEXT ,
                efficiency TEXT ,
                Scale_Table TEXT ,
                pump_wiring_connection_figure TEXT ,
                pump_figure TEXT
                )"""
        self.RunQueryNoReturn(Table)
        if basicString == True:
            self.RunQueryNoReturn(f"INSERT INTO {tableName} VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                             Values=('DefaultTestModel', ' 16 бар', ' -15C-110C', ' 40С', ' IE2', ' 3~380V/50 Hz', ' 2900 1/min', ' P2 15 кВт', ' 27 А', ' IP54', ' F', ' DN100', ' DN100', ' 480 мм', ' GG20', ' GG20', ' 1.4021', ' 335 ', '[([30.0, 101.92, 124.76], [19.96, 16.98, 9.65]), ([27.0, 98.92, 121.76], [15.96, 12.98, 5.65]), ([32.0, 101.92, 124.76], [21.96, 16.98, 9.65])]', '[[0.0, 11.77, 23.54, 35.31], [1.14, 1.17, 1.2, 1.24]]', '[[0.0, 11.77, 23.54, 35.31], [3.57, 3.84, 4.21, 4.67]]', '[[5.884298443, 11.76859689, 17.65289533, 23.53719377], [18.68826521, 19.21745255, 19.86845215, 20.71937448]]', "{'l': '34'}", 'connection_figures/test_connection.png', 'pump_figures/Test_pump_image.png'))




    # получает два имени таблицы - 1 таблица содержит 2 колонки в первой имена колонок во второй таблице и во второй колонке тип имени.
    # получает новый список колонок . 1 таблица изменяется в соответствии с списком и приводит колонки второй таблицы в соответствии с первой
    def editColumnsAccordingTo(self,TableName1, TableName2, NewColsList):
        if type(NewColsList) == str:
            try:
                NewColsList = ast.literal_eval(NewColsList)
            except Exception as e:
                print(e)
                pass
        def GiveColNames():
            Col1Names = self.RunQuery(f"SELECT Column_Name FROM {TableName1}")
            Col1Names = [each[0] for each in Col1Names]
            Col2Names = self.giveColumnDescription(TableName2, what="Names")
            return Col1Names, Col2Names
        Col1Names, Col2Names = GiveColNames()
        print("Col1Names",Col1Names)
        print("Col2Names",Col2Names)
        print("NewColsList",NewColsList)

        #операция приводит таблицу 1 (описания таблицы 2) в соответствии с новым списком
        #добавляем новое
        for newColumnDescription in NewColsList:
            if newColumnDescription[0] not in Col1Names:
                self.RunQueryNoReturn(f"INSERT INTO {TableName1} VALUES(?,?,?)", Values=(newColumnDescription[0],newColumnDescription[1],newColumnDescription[2]))
        # Удаляем чего нет
        NewColumns = [each[0] for each in NewColsList]
        print("NewColumns",NewColumns)#должен быть список
        for ExistingColumn in Col1Names:
            if ExistingColumn not in NewColumns:
                DeleteQuery = f"DELETE FROM {TableName1} WHERE Column_Name ='{ExistingColumn}' "
                # print(DeleteQuery)
                self.RunQueryNoReturn(DeleteQuery)

        #операция приводит колонки второй таблицы в соответствии записями в первой таблице
        #добавляем новое
        Col1Names, Col2Names = GiveColNames()
        for columnName in Col1Names:
            if columnName not in Col2Names:
                self.RunQueryNoReturn(f"ALTER TABLE {TableName2} ADD {columnName} TEXT")
        # Удаляем чего нет
        for ExistingColumn in Col2Names:
            if ExistingColumn not in Col1Names:
                # print("TableName2",TableName2)
                # print("ExistingColumn",ExistingColumn)
                quary = f"ALTER TABLE {TableName2} DROP COLUMN {ExistingColumn};"
                # print(quary)
                self.RunQueryNoReturn(quary)
                # sqlite до 3.35.0 не умеет удалять колонки, поэтому надо создать копию таблицы, удалить оригинальную, создать заново и поколоночно добавить из копии все столбцы, кроме удаляемых
                pass

    def __init__(self, DBName):
        self.DBName = DBName


# в бд в итоге три таблицы.
# 1 - таблица в одну строку и хранит ключ админа
# 2 - таблица со списком колонок третьей и к каждой колонке соответствует её тип
# 3 - непосредственно данные насосов.



