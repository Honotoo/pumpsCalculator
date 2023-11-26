# вот этот блок сам поставит библиотеки из списка
def install_and_import(package):
    try:
        import importlib
        importlib.import_module(package)
        print("importlib установил: ", package)
    except ImportError:
        try:
            import subprocess
            import sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
            print("subprocess установил: ", package)
        except:
            try:
                import pip
                pip.main(['install', package])
                print("pip установил: ", package)
            except Exception as e:
                print("ни один способ импорта на сработал. Ошибка: "+str(e))
    finally:
        try:
            globals()[package] = importlib.import_module(package)
        except Exception as e:
            print("не получается напрямую подключить библиотеку ",package)
            print("ошибка: ", e)
            pass

usingList = ['openpyxl','sklearn','numpy','matplotlib','scipy','sqlite3','fpdf','pillow']

for module in usingList:
    install_and_import(module)





import tkinter
from tkinter import ttk
import re
import os
from fpdf import FPDF
import os
from PIL import Image
import time
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import fsolve
from mpl_toolkits.axisartist.parasite_axes import HostAxes, ParasiteAxes
from scipy import optimize
import re
import sqlite3
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import ast


class DB_manager:
    def RunQueryNoReturn(self, SQLQuery, Values=None, maxRetry=1):
        not_done_flag = False
        maxRetry = maxRetry
        con = sqlite3.connect(self.DBName)
        Cursor = con.cursor()
        if Values != None:
            Cursor.execute(SQLQuery, Values)
        else:
            Cursor.execute(SQLQuery)
        con.commit()
        Cursor.close()
        con.close()

    def RunQuery(self, SQLQuery, maxRetry=1):
        not_done_flag = False
        maxRetry = maxRetry
        con = sqlite3.connect(self.DBName)
        Cursor = con.cursor()
        Cursor.execute(SQLQuery)
        results = Cursor.fetchall()
        con.commit()
        Cursor.close()
        con.close()
        return results




    def tableWithSingleValueOperation(self, TableName,Operation, newValue = ""):
    # create, get, set, update
        if Operation == "create":
            AdminKeyTable = """CREATE TABLE IF NOT EXISTS {} (id_ INTEGER, value TEXT)""".format(TableName)
            self.RunQueryNoReturn(AdminKeyTable)
        if Operation == "set":
            try:
                DeleteQuery = f"DELETE FROM {TableName} WHERE id_ ='{0}' "
                self.RunQueryNoReturn( DeleteQuery)
            finally:
                SQLquery = f"INSERT INTO {TableName}  VALUES(?,?)"
                self.RunQueryNoReturn(SQLquery, Values=(0,str(newValue)))
            # self.updateQuery = f"UPDATE AdminKey SET booler ='{num}' WHERE id_ =0 "
            # RunQueryNoReturn(DBName,updateQuery)
        if Operation == "get":
            try:
                table = self.RunQuery("SELECT * FROM {}".format(TableName))
                return table[0][1]
            except:
                self.tableWithSingleValueOperation(TableName,"set", newValue = "")
                table = self.RunQuery("SELECT * FROM {}".format(TableName))
                return table[0][1]


        pass
    def giveColumnDescription(self, TableName, what = "Names"):
        parametr = 1
        if what == "Names":
            parametr =1
        elif what =="Types":
            parametr = 2
        String = "PRAGMA table_info({});".format(TableName)
        columnNames = []
        for columnDescription in self.RunQuery(String):
            columnNames.append(columnDescription[parametr])
        return columnNames

    def ReadXLS(self, fileName, max_columns=None, max_rows=None):
        filename = '%s.xlsx' % (fileName)
        wb = load_workbook(filename=filename, read_only=False)
        ws = wb.active
        XLSTable = []
        if max_columns == None:
            max_columns = ws.max_column
        if max_rows == None:
            max_rows = ws.max_row
        for row in ws.iter_rows(min_row=1, max_col=max_columns, max_row=max_rows):
            rowList = []
            for cell in row:
                rowList.append(cell.value)
            XLSTable.append(rowList)
        return XLSTable

    def writeXLS(self, fileName):
        titles = self.giveColumnDescription(fileName, what = "Names")
        data = self.RunQuery("SELECT * FROM {}".format(fileName))
        filename = '%s.xlsx' % (fileName)
        wb = Workbook()
        ws = wb.active
        ws.append(titles)
        wb.save(filename)
        wb.close()

        wb = load_workbook(filename=filename, read_only=False)
        ws = wb.active
        for each in data:
            ws.append(each)
        wb.save(filename)
        wb.close()
        return filename

    def CreateTableWithThisXLS(self, FileAndTableName="Pumpsxlsx"):

        XLTable = self.ReadXLS(FileAndTableName)
        # print(len(XLTable))
        for row in XLTable:
            # если строка первая, то там названия колонок, так что мы создаем таблицу и создаем эти колонки какие есть
            if XLTable.index(row) == 0:
                ColumnNamesString = ''
                for each in range(len(row)):

                    typeOfTHisColumn = " TEXT "
                    # try:
                    #     itsInt = int(XLTable[1][each])
                    #     if str(type((itsInt))) == "<class 'int'>":
                    #         typeOfTHisColumn = " INTEGER "
                    # except:
                    try:
                        itsInt = str(XLTable[1][each])
                        if str(type((itsInt))) == "<class 'str'>":
                            typeOfTHisColumn = " TEXT "
                    except Exception as e:
                        print("Ошибка DBFunctions ", e)
                        print("возмоожно  excel пустая или есть пустые подозрительные места")
                    ColumnNamesString = ColumnNamesString + " " + row[each] + typeOfTHisColumn
                    if row.index(row[each]) < len(row) - 1:
                        ColumnNamesString = ColumnNamesString + ","
                Table = """CREATE TABLE IF NOT EXISTS {} (
                                {}
                                )""".format(FileAndTableName, ColumnNamesString)
                # print(Table)
                self.RunQueryNoReturn(Table)
                continue

            # дальше мы построчно добавляем строки
            amount = ""
            for each in range(len(row)):
                amount = amount + "?"
                if row.index(row[each]) < len(row) - 1:
                    amount = amount + ","

            # print(amount)
            SQLquery = f"INSERT INTO %s VALUES(%s)" % (FileAndTableName, amount)

            Values = []

            columnTypes = self.giveColumnDescription(FileAndTableName, what = "Types")
            # print(columnTypes)
            for each in range(len(row)):
                if columnTypes[each] == 'INTEGER':
                    Values.append(int(row[each]))
                elif columnTypes[each] == 'TEXT':
                    Values.append(str(row[each]))
            print(SQLquery)
            print(Values)
            self.RunQueryNoReturn(SQLquery, Values=Values)

    def setDefaultColumnsList(self,tableName):
        Table = f"""CREATE TABLE IF NOT EXISTS {tableName}(
                Column_Name TEXT ,
                Column_Type TEXT,
                Column_Description TEXT
                )"""
        self.RunQueryNoReturn(Table)
        self.RunQueryNoReturn(f"INSERT INTO {tableName} VALUES(?,?,?)", Values=("Model_Name", "Данные о продукте", "Название модели"))


    def setDefaultBaseTable(self,tableName, basicString = True):
        Table = f"""CREATE TABLE IF NOT EXISTS {tableName}(
                Model_Name TEXT ,
                max_pressure TEXT ,
                liquid_pump_temp TEXT ,
                max_env_temp TEXT ,

                engine_efficiency_class TEXT ,
                Net_connection TEXT ,
                rotation_frequency_nominal TEXT ,
                Nominal_power_P2 TEXT ,
                Nominal_currency TEXT ,
                Moisture_protection TEXT ,
                Isolation_Class TEXT ,

                Sucking_pipe_branch TEXT ,
                Pushing_pipe_branch TEXT ,
                building_length TEXT ,

                frame TEXT ,
                working_wheel TEXT ,
                Shaft TEXT ,
                Weight TEXT ,

                lift_H_m TEXT ,
                Cavitation_NPSH_m TEXT ,
                motor_power_P2_kW TEXT ,
                efficiency TEXT ,
                Scale_Table TEXT ,
                pump_wiring_connection_figure TEXT ,
                pump_figure TEXT
                )"""
        self.RunQueryNoReturn(Table)
        if basicString == True:
            self.RunQueryNoReturn(f"INSERT INTO {tableName} VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                             Values=('DefaultTestModel', ' 16 бар', ' -15C-110C', ' 40С', ' IE2', ' 3~380V/50 Hz', ' 2900 1/min', ' P2 15 кВт', ' 27 А', ' IP54', ' F', ' DN100', ' DN100', ' 480 мм', ' GG20', ' GG20', ' 1.4021', ' 335 ', '[([30.0, 101.92, 124.76], [19.96, 16.98, 9.65]), ([27.0, 98.92, 121.76], [15.96, 12.98, 5.65]), ([32.0, 101.92, 124.76], [21.96, 16.98, 9.65])]', '[[0.0, 11.77, 23.54, 35.31], [1.14, 1.17, 1.2, 1.24]]', '[[0.0, 11.77, 23.54, 35.31], [3.57, 3.84, 4.21, 4.67]]', '[[5.884298443, 11.76859689, 17.65289533, 23.53719377], [18.68826521, 19.21745255, 19.86845215, 20.71937448]]', "{'l': '34'}", 'connection_figures/test_connection.png', 'pump_figures/Test_pump_image.png'))




    # получает два имени таблицы - 1 таблица содержит 2 колонки в первой имена колонок во второй таблице и во второй колонке тип имени.
    # получает новый список колонок . 1 таблица изменяется в соответствии с списком и приводит колонки второй таблицы в соответствии с первой
    def editColumnsAccordingTo(self,TableName1, TableName2, NewColsList):
        if type(NewColsList) == str:
            try:
                NewColsList = ast.literal_eval(NewColsList)
            except Exception as e:
                print(e)
                pass
        def GiveColNames():
            Col1Names = self.RunQuery(f"SELECT Column_Name FROM {TableName1}")
            Col1Names = [each[0] for each in Col1Names]
            Col2Names = self.giveColumnDescription(TableName2, what="Names")
            return Col1Names, Col2Names
        Col1Names, Col2Names = GiveColNames()
        print("Col1Names",Col1Names)
        print("Col2Names",Col2Names)
        print("NewColsList",NewColsList)

        #операция приводит таблицу 1 (описания таблицы 2) в соответствии с новым списком
        #добавляем новое
        for newColumnDescription in NewColsList:
            if newColumnDescription[0] not in Col1Names:
                self.RunQueryNoReturn(f"INSERT INTO {TableName1} VALUES(?,?,?)", Values=(newColumnDescription[0],newColumnDescription[1],newColumnDescription[2]))
        # Удаляем чего нет
        NewColumns = [each[0] for each in NewColsList]
        print("NewColumns",NewColumns)#должен быть список
        for ExistingColumn in Col1Names:
            if ExistingColumn not in NewColumns:
                DeleteQuery = f"DELETE FROM {TableName1} WHERE Column_Name ='{ExistingColumn}' "
                # print(DeleteQuery)
                self.RunQueryNoReturn(DeleteQuery)

        #операция приводит колонки второй таблицы в соответствии записями в первой таблице
        #добавляем новое
        Col1Names, Col2Names = GiveColNames()
        for columnName in Col1Names:
            if columnName not in Col2Names:
                self.RunQueryNoReturn(f"ALTER TABLE {TableName2} ADD {columnName} TEXT")
        # Удаляем чего нет
        for ExistingColumn in Col2Names:
            if ExistingColumn not in Col1Names:
                # print("TableName2",TableName2)
                # print("ExistingColumn",ExistingColumn)
                quary = f"ALTER TABLE {TableName2} DROP COLUMN {ExistingColumn};"
                # print(quary)
                self.RunQueryNoReturn(quary)
                # sqlite до 3.35.0 не умеет удалять колонки, поэтому надо создать копию таблицы, удалить оригинальную, создать заново и поколоночно добавить из копии все столбцы, кроме удаляемых
                pass

    def __init__(self, DBName):
        self.DBName = DBName


def split_size_table(stroka):
    try:
        if stroka == "":
            return {"size": "?"}
        Return_Dictionary = {}
        K_V_list = stroka.split(',')
        for each in K_V_list:
            Return_Dictionary[each.split(":")[0]] = each.split(":")[1]
        return Return_Dictionary
    except Exception as e:
        return {"-":"-"}





def splitter(stroka = str):
    X = []
    Y = []
    # если введенные точки явно откуда-то скопированы, то там не иначе как будет переход на другую строку
    if '\n' in stroka:
        spisok = stroka.split('\n')
        # если разделители такие, то запятая будет делить целое и не целое
        if '\t' in stroka or ';' in stroka:
            for each in spisok:
                # print(each)
                if ";" in each:
                    X.append(each.split(";")[0])
                    Y.append(each.split(";")[1])
                elif "\t" in each:
                    X.append(each.split("\t")[0])
                    Y.append(each.split("\t")[1])
            if "," in X[0]:
                X = [float(re.sub(r',', '.', element)) for element in X]
                Y = [float(re.sub(r',', '.', element)) for element in Y]
        # Если есть одновременно и запятая и точка. То запятая делит X, Y
        elif ',' in stroka and '.' in stroka:
            for each in spisok:
                try:
                    X.append(float(each.split(",")[0]))
                    Y.append(float(each.split(",")[1]))
                except:
                    pass
        elif ',' in stroka and not '.' in stroka:
            for each in spisok:
                try:
                    # print(each)
                    # print(each.split(","))
                    # X.append(float(each.split(",")[0]))
                    Y.append(float(re.sub(r',', ".", each, flags=re.DOTALL)))
                except:
                    pass
    elif ',' in stroka  and '\n' not in stroka:
        if "!" in stroka:
            xy = stroka.split("!")
            xspisok = xy[0].split(',')
            for each in xspisok:
                try:
                    X.append(float(each))
                except:
                    pass
            yspisok = xy[1].split(',')
            for each in yspisok:
                try:
                    Y.append(float(each))
                except:
                    pass
            pass
        else:
            spisok = stroka.split(',')
            for each in spisok:

                try:
                    Y.append(float(each))
                except:
                    pass
    # Если была введена жесть, то она выкидывает те пары значений, которые не цифры
    if len(X)>=1 and type(X[0]) == str:
        # while str in X:
        indexator = 0
        while indexator<len(X) or indexator<len(Y):
            try:
                X[indexator] = float(re.sub(r',', '.', X[indexator], flags=re.DOTALL))
                Y[indexator] = float(re.sub(r',', '.', Y[indexator], flags=re.DOTALL))
                indexator+=1
            except:
                X.pop(indexator)
                Y.pop(indexator)
    try:
        promezhutok = ast.literal_eval(stroka)
        if type(promezhutok) != tuple:
            if len(promezhutok) == 2:
                X = promezhutok[0]
                Y = promezhutok[1]
            else:
                Y = promezhutok
        return X, Y
    except:
        pass
    return X, Y

# splitter разделяет все вот эти варианты на два списка X, Y если они состоят из двух или только Y если там только он.


def splitterabstracted(stroka = str):
    # stroka = List_cleaner(stroka)
    # print(stroka)
    X = []
    Y = []
    # если введенные точки явно откуда-то скопированы, то там не иначе как будет переход на другую строку
    if "!" in stroka:
        xy = stroka.split("!")
        speeds = []
        for pair in xy:
            speeds.append(splitter(stroka=pair))
        return speeds
    else:
        return [splitter(stroka=stroka)]





def plotThis(PlotDict, commonEnv = True):
    # две не solid штуки, которые мы тащим со словоря
    # они по идее должны идти отдельно, но поскольку мы в программе шлем словари, а не картинки
    # для скорости.имя оси и точка пересечения упакованы в словарь

    xAxisName = "Ось X"
    try:
        xAxisName = PlotDict['xAxisName']
        PlotDict.pop('xAxisName')
    except:
        pass
    points = None
    try:
        points = PlotDict['points']
        PlotDict.pop('points')
    except:
        pass

    fig = plt.figure(figsize = [12, 6])

    host = fig.add_axes([0.15, 0.1, 0.65, 0.8], axes_class=HostAxes)
    host.axis["right"].set_visible(False)
    host.axis["left"].set_visible(False)

    d = {}
    maxX = 0
    minX = 0
    maxY = 0
    minY = 0
    for k,XY in PlotDict.items():
        X,Y = XY[0],XY[1]
        par1 = ParasiteAxes(host)
        p1, = par1.plot(X, Y, label=str(k))
        d[par1] = p1
        maxX = max(max(X),maxX)
        minX = min(min(X), minX)
        maxY = max(max(Y), maxY)
        minY = min(min(Y), minY)

    maxX = int(maxX/100*110)
    minX = int(minX/100*110)
    maxY = int(maxY/100*110)
    minY = int(minY/100*110)

    host.set_xlim(minX, maxX)
    host.set_ylim(minY, maxY)

    indexer = 0
    for AP, kXY in zip(d.items(), PlotDict.items()):
        Axeses, Plots = AP[0], AP[1]
        k, X, Y = kXY[0], kXY[1][0], kXY[1][1]
        host.parasites.append(Axeses)
        Axeses.axis["right"].set_visible(False)

        if indexer == 0:
            Axeses.axis["left"] = Axeses.new_fixed_axis(loc="left", offset=(0, 0))
            Axeses.axis["left"].label.set_color(Plots.get_color())
            if commonEnv:
                Axeses.set(ylim=(minY, maxY),xlim=(minX,maxX), ylabel=k)
            else:
                Axeses.set( ylabel=k)
            Axeses.grid(True)
        else:
            Axeses.axis["right"] = Axeses.new_fixed_axis(loc="right", offset=((indexer-1)*40, 0))
            if commonEnv:
                Axeses.set(ylim=(minY, maxY),xlim=(minX,maxX), ylabel=k)
            else:
                Axeses.set( ylabel=k)
            Axeses.axis["right"].label.set_color(Plots.get_color())

        host.set(xlabel=xAxisName)
        # par1.set(ylim=(0, 4), ylabel="Temperature")
        indexer+=1
    if points !=None:
        for point in points:

            print("point")
            print(point)
            if point[1] == 0:
                host.plot([point[0], point[0]], [0, maxY], "k--")
                host.text(point[0] + 2, 0.5, fr"{point[2]} {point[0]:.2f}", color="k", fontsize=10)

            elif point[0] == 0:
                host.plot([0, maxX], [point[1], point[1]], "k--")
                host.text(2, point[1] - 3, fr"{point[3]} {point[1]:.2f}", color="k", fontsize=10)
            else:
                host.plot([0, point[0]], [point[1], point[1]], "k--")
                host.text(2, point[1] - 3, fr"{point[3]} {point[1]:.2f}", color="k", fontsize=10)

                host.plot([point[0], point[0]], [0, point[1]], "k--")
                host.text(point[0] + 2, 0.5, fr"{point[2]} {point[0]:.2f}", color="k", fontsize=10)

                host.plot([point[0]], [point[1]], 'o', markersize=5, color="k")
                #host.plot(point[0], calculatedPoints, label="calculated")

    host.legend()
    host.grid(True)
    host.margins(x=0.01)
    # plt.show()
    return plt.gcf()  # .savefig('im.png', bbox_inches='tight')

class linModel:
    a = None
    b = None
    def linear_fit(self, x, a, b):
        return a * x + b
    def fit(self, XsofLine, YsofLine):
        self.a, self.b = optimize.curve_fit(self.linear_fit, XsofLine, YsofLine)[0]
    def predict(self, x):
        return self.linear_fit(x, self.a, self.b)
    def __str__(self):
        return "a: "+str(self.a)+" b: "+str(self.b)
    def __init__(self):
        pass


class calcManager:
    CalcValues = None
    dbList = None
    OkDBList = []
    modelType = "poly1d"#linear
    ShowAll = False
    model = None

    def plotThis(self, PlotDict):
        plotThis(PlotDict, commonEnv=True)

    def Lines_intersects(self, s0, s1):
        dx0 = s0[1][0] - s0[0][0]
        dx1 = s1[1][0] - s1[0][0]
        dy0 = s0[1][1] - s0[0][1]
        dy1 = s1[1][1] - s1[0][1]
        p0 = dy1 * (s1[1][0] - s0[0][0]) - dx1 * (s1[1][1] - s0[0][1])
        p1 = dy1 * (s1[1][0] - s0[1][0]) - dx1 * (s1[1][1] - s0[1][1])
        p2 = dy0 * (s0[1][0] - s1[0][0]) - dx0 * (s0[1][1] - s1[0][1])
        p3 = dy0 * (s0[1][0] - s1[1][0]) - dx0 * (s0[1][1] - s1[1][1])
        return (p0 * p1 <= 0) & (p2 * p3 <= 0)
    # print(Lines_intersects([(2,1),(9,5)],[(5,1),(3,5)]))
    # print(Lines_intersects([(2,1),(9,5)],[(10,0),(6,2)]))

    def complex_intersection(self, Y1, X1, Y2, X2):
        # X1 = X1[::-1], Y1=Y1[::-1], X2=X2[::-1], Y2=Y2[::-1],
        for each1 in range(len(X1) - 1):
            line1 = [(X1[each1], Y1[each1]), (X1[each1 + 1], Y1[each1 + 1])]
            for each2 in range(len(X2) - 1):
                line2 = [(X2[each2], Y2[each2]), (X2[each2 + 1], Y2[each2 + 1])]
                if self.Lines_intersects(line1, line2) == True:
                    # print("complex_intersection: "+str(line1) + str(line2))
                    return (line1, line2)
                    # return str(line1)+str(line2)
        return None

    def give_Y_of_line_at_this_X(self, Line_X1, Line_Y1, this_X):
        X_collision = 0
        try:
            dots = self.complex_intersection([0, 1000],[this_X, this_X], Line_Y1,  Line_X1)
            XsofLine = np.array([dots[1][0][0], dots[1][1][0]])
            YsofLine = np.array([dots[1][0][1], dots[1][1][1]])
            self.model = linModel()
            self.model.fit(YsofLine, XsofLine)
            X_collision = fsolve(lambda k: this_X - self.model.predict(k), 42)[0]
        except:
            pass
        return X_collision

    def calculator(self, ExistingLineX, ExistingLineY, Q, H, percents, modelType = "linear", checkInterceptBool = False):

        # формула непосредственно для расчета кривой новой
        def Formula(Q, H, EfficiancyValue):
            return (EfficiancyValue /Q) ** 2 *H
        # для новой кривой немного сдвигается значение X по вот этой логике
        def FormulaX (Q, percent_):
            return percent_/100*Q
        # делаем  Y для точек в новой кривой
        def calculate_Hvx_Line_Y(Q, H, Y):
            calculatedPoints = []
            for each in Y:
                calculatedPoints.append(Formula(Q, H, each))
            return calculatedPoints
        # делаем  X для точек в новой кривой
        def calculate_Hvx_Line_X(Q, un_destorted_list):
            to_return=[]
            for percent_ in un_destorted_list:
                to_return.append(FormulaX (Q, percent_))
            return to_return

        # график по расчету
        calculatedPoints_X = calculate_Hvx_Line_X(Q, range(1,percents+1))
        calculatedPoints_Y = np.array(calculate_Hvx_Line_Y(Q, H, calculatedPoints_X))

        # по сути, если нам нужно только проверить факт пересечения, то дальше мы можем не считать.
        if checkInterceptBool == True:
            dots = self.complex_intersection(calculatedPoints_X, calculatedPoints_Y, ExistingLineX, ExistingLineY)
            if dots == None:
                return False
            else:
                return True


        #Строим модель полиномиальной/ линейной регрессии имеющейся линии
        self.model = None
        if modelType == "poly1d":
            self.model = np.poly1d (np.polyfit (ExistingLineX, ExistingLineY , 6))
        if modelType == "linear":
            dots = self.complex_intersection(calculatedPoints_X, calculatedPoints_Y, ExistingLineX, ExistingLineY)
            XsofLine = np.array([dots[1][0][0],dots[1][1][0]])
            YsofLine = np.array([dots[1][0][1],dots[1][1][1]])
            self.model = linModel()
            self.model.fit( YsofLine,XsofLine)

        # Ищем x точки соприкосновения двух моделей
        X_collision = 0
        Y_collision = 0
        # сделаем ещё график для модели, описывающей имеющиеся данные
        X_plotModel = []
        Y_plotModel = []


        #подбор значения при котором разница X между нашей моделью по данным и функцией расчетов будет минимальной
        if modelType == "poly1d":
            X_collision = fsolve(lambda k: Formula(Q, H, k) - self.model(k), 42)[0]
            # просто вставвляем найденный икс в функцию и получаем Y по стандартной матеше
            Y_collision = Formula(Q, H, X_collision)
            for each in range(0, int(max(ExistingLineX))):
                X_plotModel.append(each)
                Y_plotModel.append(self.model(each))

        if modelType == "linear":
            X_collision = fsolve(lambda k: Formula(Q, H, k) - self.model.predict(k), 42)[0]
            # просто вставвляем найденный икс в функцию и получаем Y по стандартной матеше
            Y_collision = Formula(Q, H, X_collision)

            for each in range(0, int(max(ExistingLineX))):
                X_plotModel.append(each)
                Y_plotModel.append(self.model.predict(each))


        return  calculatedPoints_X, calculatedPoints_Y, X_collision, Y_collision,X_plotModel, Y_plotModel

    def iterating_matcher(self, showCalculatedModel = False):#, modelType = "linear"  "poly1d"#linear

        ourX = "flow_rate_Q_m3_h"
        ourY = "lift_H_m"

        def GiveCalcData():
            Q = int(self.CalcValues["Flow"])
            H = int(self.CalcValues["pressure"])
            percents = int(self.CalcValues["percents"])
            modelType = self.modelType
            return Q, H, percents, modelType

        def isThisPumpOK(pump):
            speed = 1
            for pumpSpeedLine in ast.literal_eval(pump[ourY]):
                ExistingLineX = pumpSpeedLine[0]
                ExistingLineY = pumpSpeedLine[1]
                Q, H, percents, modelType = GiveCalcData()
                # если пересечения нет, то ничего не делаем. Функция вернет false (небольшая надстройка, но мне так удобнее вышло)
                if False == self.calculator( ExistingLineX, ExistingLineY, Q, H, percents, modelType, checkInterceptBool = True):
                    speed+=1
                    continue
                else:
                    return  True, speed
            return False, 0

        def putThisPumpinOkDBList(pump,speed):
            Q, H, percents, modelType = GiveCalcData()
            currentSpeed = 1
            THisPump_X_collision, THisPump_Y_collision = None,None
            PlotDict = {}

            for pumpSpeedLine in ast.literal_eval(pump[ourY]):
                ExistingLineX = pumpSpeedLine[0]
                ExistingLineY = pumpSpeedLine[1]
                calculatedPoints_X, calculatedPoints_Y, X_collision, Y_collision, X_plotModel, Y_plotModel = self.calculator( ExistingLineX, ExistingLineY, Q, H, percents, modelType)
                if currentSpeed == speed:
                    # если нужно напечатать матмодель по данным
                    if showCalculatedModel == True:
                        PlotDict["approximated"] = [X_plotModel, Y_plotModel]
                    THisPump_X_collision, THisPump_Y_collision = X_collision, Y_collision
                    point = (X_collision, Y_collision, r'Q (m$^3$/s)', 'h (m)')
                    PlotDict['points'] = [point]
                    PlotDict['xAxisName'] = ourX
                    PlotDict["Рассчитанная модель"] = [calculatedPoints_X, list(calculatedPoints_Y)]
                PlotDict["Скорость " + str(currentSpeed)] = [ExistingLineX, ExistingLineY]
                currentSpeed+=1
            PlotDict['КПД'] = ast.literal_eval(pump['efficiency'])[0]
            pump["PlotDict1"] = PlotDict


            linesPlot2 = ['motor_power_P2_kW', 'Cavitation_NPSH_m']
            PlotDict = {}
            for lineName in linesPlot2:
                PlotDict[lineName] = ast.literal_eval(pump[lineName])[0]
                point = (THisPump_X_collision, 0, r'Q (m$^3$/s)', ' ')
                PlotDict['points'] = [point]
                PlotDict['xAxisName'] = "Ось Х"
                # plotThis(PlotDict)
            pump["PlotDict2"] = PlotDict

            # сбор данных для предварительного отчета и для вывода в Заданных параметрах
            IntersectionData = {}
            proizvNapor = pump["PlotDict1"]["points"]
            IntersectionData["Модель"] = pump["Model_Name"]
            IntersectionData["Производительность"] = f"{proizvNapor[0][0]:.2f} " + " м3/ч"
            IntersectionData["Напор"] = f"{proizvNapor[0][1]:.2f} " + " (m)"

            motor_power_P2_kW = self.give_Y_of_line_at_this_X(ast.literal_eval(pump["motor_power_P2_kW"])[0][0],
                                        ast.literal_eval(pump["motor_power_P2_kW"])[0][1], PlotDict['points'][0][0])
            IntersectionData["Мощность на валу"] = f"{motor_power_P2_kW:.2f} "#  motor_power_P2_kW

            efficiency = self.give_Y_of_line_at_this_X(ast.literal_eval(pump["efficiency"])[0][0], ast.literal_eval(pump["efficiency"])[0][1], PlotDict['points'][0][0])
            IntersectionData["КПД"] = f"{efficiency:.2f} "# efficiency

            Cavitation_NPSH_m = self.give_Y_of_line_at_this_X(ast.literal_eval(pump["Cavitation_NPSH_m"])[0][0],
                                        ast.literal_eval(pump["Cavitation_NPSH_m"])[0][1], PlotDict['points'][0][0])
            IntersectionData["NPSH"] = f"{Cavitation_NPSH_m:.2f} "

            pump["Заданные параметры"] = IntersectionData
            self.OkDBList.append(pump)

        for pump in self.dbList:
            ok, speed = isThisPumpOK(pump)
            if True == ok:
                putThisPumpinOkDBList(pump,speed)

    def giveMeGoodPumps(self, showCalculatedModel = False):
        self.OkDBList = []
        self.iterating_matcher(showCalculatedModel)

        return self.OkDBList


    def __init__(self,toCalc):
        # print( toCalc)
        self.CalcValues = toCalc["CalcValues"]
        self.dbList = toCalc["dbList"]
        self.modelType = toCalc["modelType"]
        self.ShowAll = toCalc["ShowAll"]



    pass



class pdfManager:
    columnDescriptions = None
    PumpData = None


    cursorPosition = 25
    interval = 4
    TableLeftAlign = 125
    FontSize = 10
    imgOffset = 22
    CachImafesNames = ['cache_images\H_Q_plot.png', 'cache_images\_NPSH_P2_plot.png']
    pdf = None

    def makeBasicPDF(self):
        class PDF(FPDF):
            pass
        self.pdf = PDF()
        self.pdf.alias_nb_pages()
        self.pdf.add_page()
        self.pdf.add_font('Times', '', 'TimesNewRomanRegular.ttf', uni=True)
        self.pdf.add_font('TimesB', 'B', 'Times New Roman Bold.ttf', uni=True)
        self.Font()
        return self.pdf

    def move(self):
        self.cursorPosition = self.cursorPosition + self.interval
        return self.cursorPosition

    def Font(self,bold = False, FontSize = 10 ):
        if bold == True:
            self.pdf.set_font('TimesB', style='B', size=FontSize)
        else:
            self.pdf.set_font('Times', style='', size=FontSize)

    def printLine(self,text1, text2, Bold = False):
        if Bold:
            self.Font(bold = True, FontSize = 10)
        self.pdf.text(self.TableLeftAlign, self.cursorPosition, text1)
        self.pdf.text(self.TableLeftAlign + 53, self.cursorPosition, text2)
        self.Font()
        self.move()

    def tablePart(self,Header, CategoryAndValue):
        self.move()
        self.printLine(Header, "", Bold = True)
        for Category, Value in CategoryAndValue.items():
            self.printLine(Category, Value)
        # return cursorPosition

    def assembleTable(self):
        resultingTable = {}
        # собрать пары ключ - имя колонки значение - заголовок колонки
        for Column_Description in self.columnDescriptions:
            if Column_Description[1] != 'не текст' and Column_Description[1] not in resultingTable.keys():
                resultingTable[Column_Description[1]] = {}
        # собрать словарь ключ - заголовок группы. Ключ - словарь из описания поля и значения поля
        for each in self.columnDescriptions:
            if each[1] in resultingTable.keys():
                resultingTable[each[1]][each[2]] = self.PumpData[each[0]]
        # фильтруем поля. Если есть такие, в которых есть только (None, "", " ")  то их выкинем

        resultingTable2 = {}
        for groupName, group in resultingTable.items():
            resultingTable2[groupName] = {}
            deleteGroup = True
            for itemName, itemValue in group.items():
                if itemValue not in [None, "", " ","None"]:
                    resultingTable2[groupName][itemName] =itemValue
                    deleteGroup = False
            if deleteGroup == True:
                resultingTable2.pop(groupName)
        return resultingTable2
        pass


    def givePDF(self):
        self.makeBasicPDF()
        print(self.PumpData)

        self.tablePart("Заданные параметры", self.PumpData["Заданные параметры"])

        assembledTable = self.assembleTable()
        for Header, CategoryAndValue in assembledTable.items():
            self.tablePart(Header, CategoryAndValue)

        name1 = 'cache_images\H_Q_plot.png'
        name2 = 'cache_images\_NPSH_P2_plot.png'
        img1 = plotThis(self.PumpData['PlotDict1'], commonEnv=True)
        img1.savefig(name1, bbox_inches='tight')
        img2 = plotThis(self.PumpData['PlotDict2'], commonEnv=True)
        img2.savefig(name2, bbox_inches='tight',transparent = True)

        width1, height1 = img1.get_size_inches() * img1.get_dpi()
        width2, height2 = img2.get_size_inches() * img2.get_dpi()
        pump_im = Image.open(self.PumpData["pump_figure"])
        width3, height3 = pump_im.size


        self.pdf.image(name1, x=10, y=self.imgOffset, w=100)
        self.pdf.image(name2, x=10, y=self.imgOffset + height1 / 10 + 2, w=100)
        self.pdf.image(self.PumpData["pump_figure"], x=10, y=self.imgOffset + 4 + height1 / 10 + height2 / 10 + 4, w=100)

        TheTableData = ast.literal_eval(self.PumpData["Scale_Table"])

        self.pdf.set_xy(10, self.imgOffset + 4 + height1 / 10 + height2 / 10 + height3 / 11)
        for k, v in TheTableData.items():
            self.pdf.cell(12, 6, str(k), 1, 0, 'C')

        self.pdf.set_xy(10, self.imgOffset + 4 + height1 / 10 + height2 / 10 + height3 / 11 + 6)
        for k, v in TheTableData.items():
            self.pdf.cell(12, 6, str(v), 1, 0, 'C')


        self.pdf.set_xy(self.TableLeftAlign - 2, self.cursorPosition)
        self.pdf.cell(40, 6, "Мощность двигателя P2 <= 3 кВт", 0, 0, 'C')
        self.pdf.cell(40, 6, "        P2 >= 3 кВт", 0, 0, 'C')
        self.pdf.image(self.PumpData['pump_wiring_connection_figure'], x=self.TableLeftAlign, y=self.cursorPosition + 6, w=80)

        Filename = str("pdff.pdf")
        try:
            Filename = self.PumpData["Model_Name"]+ ".pdf"
        except:
            pass

        print(Filename)
        self.pdf.output(Filename, 'F')
        try:
            os.startfile(Filename)
        except:
            os.system("xdg-open \"%s\"" % Filename)

    def __init__(self,columnDescriptions,PumpData):
        self.columnDescriptions = columnDescriptions
        self.PumpData = PumpData






class fieldAndLabel:
    myName = "UnnamedfieldAndLabel"
    label = ""
    raw = 0
    column = 0
    entryDefault = ''
    entry = None#tkinter.Entry()

    def set(self, value, alsoDefault = False):
        self.entry.delete(0, len(self.entry.get()))
        self.entry.insert(0, str(value))
        if alsoDefault == True:
            self.entryDefault = str(value)
    def content(self):
        try:
            return self.entry.get()
        except:
            print("объект "+self.myName+ " не может взять данные из поля")
            return self.myName,None
    def clear(self):
        try:
            self.entry.delete(0, len(self.entry.get()))
            self.entry.insert(0, self.entryDefault)
        except Exception as e:
            print("не получается очистить Entry в  " + self.myName)
        pass

    def __init__(self, frame, myName, Label, raw, column, entryDefault="" ):
        self.myName = myName
        self.label = Label
        self.raw = raw
        self.column = column
        self.entryDefault = entryDefault

        tkinter.Label(frame, text=self.label).grid(row=self.raw, column=self.column)
        self.entry = tkinter.Entry(frame, width=40)
        self.entry.grid(row=self.raw+1, column=self.column, columnspan=1)
        self.entry.insert(0, self.entryDefault)



class adminWindow:
    columnDescriptions = []
    EntriesList = {}
    db = None

    def fillContent(self,frame):
        def CountXY(amount, Xequals=1, Yequals=10):
            x = 0
            y = 0
            for Ys in range(1, 100):
                for Xs in range(1, 100):
                    if Ys * Yequals * Xs * Xequals >= amount:
                        y = Ys * Yequals
                        x = Xs * Xequals
                        return x, y
            return x, y
        def count(xNow,yNow):
            yNow+=1
            if yNow>X:
                yNow=0
                xNow+=1
            return xNow, yNow
        xNow,yNow = 0,0

        additionalPlacesToDB = 20
        X, Y = CountXY(len(self.columnDescriptions)+additionalPlacesToDB, Xequals=1, Yequals=4)

        flow_rate_Q_m3_h = fieldAndLabel(frame, "flow_rate_Q_m3_h", "Расход - Q(m3 / h) (исп. если не введен для H/NPSH/P2/КПД)", yNow * 2, xNow, entryDefault="")
        self.EntriesList["flow_rate_Q_m3_h"] = flow_rate_Q_m3_h
        xNow, yNow = count(xNow, yNow)

        for columnIndex in range(len(self.columnDescriptions)):
            entryObj = fieldAndLabel(frame, self.columnDescriptions[columnIndex][0], self.columnDescriptions[columnIndex][2], yNow*2, xNow, entryDefault="")
            self.EntriesList[self.columnDescriptions[columnIndex][0]] = entryObj
            xNow, yNow = count(xNow, yNow)

        disAssemble_line = fieldAndLabel(frame, "disAssemble_line", "разбить введенные данные в поля автоматически", yNow * 2, xNow, entryDefault="")
        self.EntriesList["disAssemble_line"] = disAssemble_line
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text="разобрать по формам", command=lambda: self.disAssemble_line(text=self.EntriesList["disAssemble_line"].content())).grid(row=yNow * 2, column=xNow, columnspan=1, rowspan=2)
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text=" Предпросмотр PDF ",command=lambda: self.ShowPDF()).grid(row=yNow * 2, column=xNow, columnspan=1, rowspan = 2)
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text=" Добавить запись в базу",command=lambda: self.add_to_DB()).grid(row=yNow * 2, column=xNow, columnspan=1, rowspan = 2)
        xNow, yNow = count(xNow, yNow)
        AdminKeyEntry = fieldAndLabel(frame,"AdminKeyEntry", "Ключ администратора", yNow * 2, xNow, entryDefault="")
        self.EntriesList["AdminKeyEntry"] = AdminKeyEntry
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text=" Сменить ключ администратора",command=lambda: self.updateAdminKey()).grid(row=yNow * 2, column=xNow, columnspan=1, rowspan = 2)
        xNow, yNow = count(xNow, yNow)
        Unload_from_DB_to_XLSX_entry = fieldAndLabel(frame,"Unload_from_DB_to_XLSX_entry", "Вывести таблицу из базы данных в excel", yNow * 2, xNow, entryDefault="PumpsColums")
        self.EntriesList["Unload_from_DB_to_XLSX_entry"] = Unload_from_DB_to_XLSX_entry
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text=" DB => XL",command=lambda: self.Unload_from_DB_to_XLSX()).grid(row=yNow * 2, column=xNow, columnspan=1, rowspan = 2)
        xNow, yNow = count(xNow, yNow)
        Upload_from_XLSX_to_DB_entry = fieldAndLabel(frame,"Upload_from_XLSX_to_DB_entry", "Занести таблицу в БД", yNow * 2, xNow, entryDefault="PumpsColums")
        self.EntriesList["Upload_from_XLSX_to_DB_entry"] = Upload_from_XLSX_to_DB_entry
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text=" XL => DB",command=lambda: self.Upload_from_XLSX_to_DB()).grid(row=yNow * 2, column=xNow, columnspan=1, rowspan = 2)
        xNow, yNow = count(xNow, yNow)
        tkinter.Button(frame, text="Привести колонки таблицы насосов\nв соответствии с таблицей описания колонок", command=lambda: self.editColumnsAccordingTo()).grid(row=yNow * 2,
                                                                                                    column=xNow,
                                                                                                    columnspan=1,
                                                                                                    rowspan=2)

        try:
            self.EntriesList["pump_figure"].set('pump_figures/Test_pump_image.png', alsoDefault = True)
        except:pass
        try:
            self.EntriesList["pump_wiring_connection_figure"].set('connection_figures/test_connection.png', alsoDefault = True)
        except:pass
        pass


    #Инструменты работы внутри окна
    def Clear_all_entry(self):
        pass
    def CollectAllForms(self):
        dict_to_row = {}
        for k,v in self.EntriesList.items():
            dict_to_row[k] = self.EntriesList[k].content()
        # lift_H_m
        dict_to_row['lift_H_m'] =str(splitterabstracted(dict_to_row['lift_H_m']))
        # Cavitation_NPSH_m
        dict_to_row['Cavitation_NPSH_m'] = str(splitterabstracted(dict_to_row['Cavitation_NPSH_m']))
        # motor_power_P2_kW
        dict_to_row['motor_power_P2_kW'] = str(splitterabstracted(dict_to_row['motor_power_P2_kW']))
        # efficiency
        dict_to_row['efficiency'] = str(splitterabstracted(dict_to_row['efficiency']))
        dict_to_row['Scale_Table'] = str(split_size_table(dict_to_row['Scale_Table']))

        return dict_to_row

        # from tkkk import alert_window
        # alert_window("не получилось собрать данные из полей. \nскорее всего неправильно введеный Расход или эффективность:\n" + str(e))
        pass

    def disAssemble_line(self, text="42"):
        variants = {"max_pressure":r"(?<=давление).*?(?=\n)",
                    "liquid_pump_temp":r"(?<=жидкости).*?(?=\n)",
                    "max_env_temp":r"(?<=Среды).*?(?=\n)",
                    "engine_efficiency_class":r"(?<=двигателя).*?(?=\n)",
                    "Net_connection":r"(?<=сети).*?(?=\n)",
                    'rotation_frequency_nominal':r"(?<=вращения).*?(?=\n)",
                    "Nominal_power_P2":r"(?<=мощность).*?(?=\n)",
                    "Nominal_currency":r"(?<=ток).*?(?=\n)",
                    "Moisture_protection":r"(?<=защиты).*?(?=\n)",
                    "Isolation_Class":r"(?<=изоляции).*?(?=\n)",
                    "Sucking_pipe_branch":r"(?<=всасывания).*?(?=\n)",
                    "Pushing_pipe_branch":r"(?<=напорной стороне).*?(?=\n)",
                    "building_length":r"(?<=длина).*?(?=\n)",
                    "frame":r"(?<=Корпус).*?(?=\n)",
                    "working_wheel":r"(?<=колесо).*?(?=\n)",
                    "Shaft":r"(?<=Вал).*?(?=\n)",
                    "Weight":r"(?<=Вес).*?(?=кг)"
                    }
        for k,v in variants.items():
            try:
                self.EntriesList[k].set(re.search(variants[k], text)[0])
            except:
                pass


    def ShowPDF(self):
        The_Line_dict = self.CollectAllForms()
        print(The_Line_dict)

    def add_to_DB(self):
        # собираем контент с полей
        The_Line_dict =  self.CollectAllForms()
        Columns_ordered = self.db.giveColumnDescription("Pumps", what = "Names")
        print(The_Line_dict)
        NewLineOrdered = []
        for each in Columns_ordered:
            NewLineOrdered.append(The_Line_dict[each])


        amount = ""
        for each in range(len(NewLineOrdered)):
            amount = amount + "?"
            if NewLineOrdered.index(NewLineOrdered[each]) < len(NewLineOrdered) - 1:
                amount = amount + ","

        SQLquery = f"INSERT INTO Pumps VALUES({amount})"
        print(SQLquery)
        self.db.RunQueryNoReturn(SQLquery, Values = NewLineOrdered)
        self.Clear_all_entry()

    def updateAdminKey(self):
        num = self.EntriesList["AdminKeyEntry"].content()
        self.db.tableWithSingleValueOperation("AdminKey","set", newValue = "num")
        self.EntriesList["AdminKeyEntry"].clear()
        print(num)

    def Upload_from_XLSX_to_DB(self):
        try:
            FileName = self.EntriesList["Upload_from_XLSX_to_DB_entry"].content()
            self.EntriesList["Upload_from_XLSX_to_DB_entry"].clear()
            try:
                self.db.RunQueryNoReturn(f"DROP TABLE {FileName};")
            except:
                pass
            self.db.CreateTableWithThisXLS(FileAndTableName=FileName)
            pass
        except Exception as e:
            print(e)

    def Unload_from_DB_to_XLSX(self):
        try:
            FileName = self.EntriesList["Unload_from_DB_to_XLSX_entry"].content()
            self.EntriesList["Unload_from_DB_to_XLSX_entry"].clear()
            FileName = self.db.writeXLS(FileName)
            try:
                os.startfile(FileName)
            except:
                os.system("xdg-open \"%s\"" % FileName)
            pass
        except Exception as e:
            print(e)
            # from tkkk import alert_window
            #alert_window("выгрузить базу данных в XLS не вышло:\n" + str(e))
    def editColumnsAccordingTo(self):
        ndb = self.db.RunQuery("SELECT * FROM {}".format("PumpsColums"))
        self.db.editColumnsAccordingTo("PumpsColums", "Pumps", ndb)

    def __init__(self, dbObject):
        self.db = dbObject
        try:
            self.columnDescriptions = self.db.RunQuery("SELECT * FROM {}".format("PumpsColums"))
        except:
            self.db.setDefaultColumnsList("PumpsColums")
            self.db.setDefaultBaseTable("Pumps")
            self.columnDescriptions = self.db.RunQuery("SELECT * FROM {}".format("PumpsColums"))
        root2 = tkinter.Toplevel()
        frame = ttk.Frame(root2, padding=0)
        frame.grid()
        self.fillContent(frame)



class  singleResultPresent:
    columnDescriptions = None
    pumpUnit = None
    rowCount = 0

    def GenPdf(self):
        pdf = pdfManager(self.columnDescriptions, self.pumpUnit)
        pdf.givePDF()
        print("GenPdf")
        pass

    def putText(self,frame):
        self.rowCount = 0
        for sign, value in self.pumpUnit['Заданные параметры'].items():
            tkinter.Label(frame, text=sign)
            tkinter.Label(frame, text=value)
            self.rowCount+=1


        bLook2 = tkinter.Button(frame, text="\n Скачать отчет \n", command=lambda : self.GenPdf())
        print(self.pumpUnit)
        pass

    def __init__(self,frame, columnDescriptions, pumpUnit):
        self.columnDescriptions = columnDescriptions
        self.pumpUnit = pumpUnit

        self.putText(frame)






class calcWindow :
    frame = None
    columnDescriptions = None
    goodPumps = None
    resultUnits = []

    def fillcalcContent(self):
        for pump in self.goodPumps:
            ViewUnit = singleResultPresent(self.frame, self.columnDescriptions, pump)
            self.resultUnits.append(ViewUnit)
        pass


    def __init__(self, columnDescriptions, goodPumps):
        self.columnDescriptions = columnDescriptions
        self.goodPumps = goodPumps

        root3 = tkinter.Toplevel()



class item:
    rowCount = 0

    def GenPdf(self):
        pdf = pdfManager(self.columnDescriptions, self.pumpUnit)
        pdf.givePDF()
        print("GenPdf")
        pass
    def __init__(self,root, frm,canv,Y_position_pointer,columnDescriptions, pumpUnit):
        self.columnDescriptions = columnDescriptions
        self.pumpUnit = pumpUnit


        for sign, value in self.pumpUnit['Заданные параметры'].items():

            tkinter.Label(frm, text=sign).grid(row=self.rowCount, column=0, columnspan=2)
            tkinter.Label(frm, text=value).grid(row=self.rowCount, column=2, columnspan=1)
            self.rowCount+=1

        basicGap = 10
        bLook2 = tkinter.Button(frm, text="\n Скачать отчет \n", command=lambda: self.GenPdf())

        bLook2 = bLook2.grid(row=3, column=5, rowspan=2)
        canv.create_window(basicGap, Y_position_pointer + basicGap, anchor=tkinter.NW, window=frm)
        Y_position_pointer = Y_position_pointer + basicGap + 180

        frm = tkinter.Frame(root, width=32, height=123, bg="#ffffff", bd=2)
        frm.config(relief=tkinter.SUNKEN)
        print(pumpUnit)
        img = plotThis(pumpUnit["PlotDict1"])

        img_path = "cache_images" + "/" + pumpUnit['Model_Name'] + ".png"
        img.savefig(img_path, bbox_inches='tight')

        img = tkinter.PhotoImage(file=img_path)


        button1 = tkinter.Button(frm, image=img)
        button1.image = img
        button1.pack()
        canv.create_window(basicGap, Y_position_pointer, anchor=tkinter.NW, window=frm)



class calcW:

    def __init__(self,columnDescriptions, goodPumps):
        basicGap = 10
        TextLength = 150
        commonWidth = 456
        ImageLength = 320
        ScrollLength = len(goodPumps)*1000

        root3 = tkinter.Toplevel()  
        parent = root3
        color = 'white'
        canv = tkinter.Canvas(root3, bg=color, relief=tkinter.SUNKEN)
        canv.config(width=1200, height=1000)

        canv.config(scrollregion=(-20, 0, commonWidth, ScrollLength))
        canv.config(highlightthickness=6)

        ybar = tkinter.Scrollbar(root3)
        ybar.config(command=canv.yview)
        canv.config(yscrollcommand=ybar.set)
        ybar.pack(side=tkinter.RIGHT, fill=tkinter.Y)
        canv.pack(side=tkinter.LEFT, expand=tkinter.NO, fill=tkinter.BOTH)

        Y_position_pointer = 0
        spisok_k_frm = []
        for ctr in range(len(goodPumps)):
            frm = tkinter.Frame(root3, width=commonWidth, height=TextLength, bg="#ffffff", bd=2)
            frm.config(relief=tkinter.SUNKEN)
            spisok_k_frm.append(frm)

        for frm,pumpUnit in zip(spisok_k_frm,goodPumps):
            i = item(root3, frm, canv, Y_position_pointer,columnDescriptions, pumpUnit)
            Y_position_pointer+=800





class mainWindow:
    db = DB_manager("DB.sql")
    CalcEntries = {}
    columnDescriptions = []
    modelType = None
    showCalculatedModel = None

    # в калькуляторе я беру из бд сразу всю таблицу и работаю сразу со всей таблицей, что довольно нагруженно,
    # но с другой стороны менеджер калькулятора сразу возврващает готовый для построения ПДФ словарь
    # в общем тут в идеале бы сделать другую систему, или бахнуть генератором, но может и не надо
    def OpenCalculator(self):

        modeltype = "linear"
        if self.modelType.get() == 0:
            modeltype = "linear"
        else:
            modeltype ="poly1d"


        showCalculated = True
        if self.showCalculatedModel.get() == 0:
            showCalculated = False
        else:
            showCalculated =True

        print(" Сбор данных для расчета")
        CalcValues = {}
        for k,v in self.CalcEntries.items():
            CalcValues[k] = v.content()
        CalcValues.pop("key")

        dbTable = self.db.RunQuery("SELECT * FROM {}".format("Pumps"))
        dbNames = self.db.giveColumnDescription("Pumps", what="Names")

        dbList = []
        for row in dbTable:
            LineDict = {}
            for name, value in zip(dbNames, row):
                LineDict[name] = value
            dbList.append(LineDict)

        toCalc = {"CalcValues":CalcValues, "modelType":modeltype, "ShowAll":False, "dbList":dbList}
        print("Подбор вариантов")
        self.columnDescriptions = self.db.RunQuery("SELECT * FROM {}".format("PumpsColums"))

        cm = calcManager(toCalc)
        goodPumps = cm.giveMeGoodPumps(showCalculatedModel = showCalculated)

        CW = calcW(self.columnDescriptions, goodPumps)





    def fillContent(self, frame):
        Flow = fieldAndLabel(frame, "Flow" , "Производительность(Q)", 0, 0, entryDefault = "100")
        pressure = fieldAndLabel(frame, "pressure" , "Напор(H)", 2, 0, entryDefault = "19")
        percents = fieldAndLabel(frame, "percents", "Проценты", 4, 0, entryDefault="120")
        key = fieldAndLabel(frame, "key", "Ключ администратора", 14, 0)

        self.CalcEntries["Flow"] = Flow
        self.CalcEntries["pressure"] = pressure
        self.CalcEntries["percents"] = percents
        self.modelType = tkinter.IntVar()
        ttk.Radiobutton( frame, text="линейная", value=0, variable=self.modelType ).grid(row=6, column=0)
        ttk.Radiobutton( frame, text="полиномиальная модель", value=1, variable=self.modelType ).grid(row=7, column=0)

        self.showCalculatedModel = tkinter.IntVar()
        ttk.Radiobutton( frame, text="Скрыть модель", value=0, variable=self.showCalculatedModel ).grid(row=8, column=0)
        ttk.Radiobutton( frame, text="Показать модель расчетов", value=1, variable=self.showCalculatedModel ).grid(row=9, column=0)

        self.CalcEntries["key"] = key
        tkinter.Button(frame, text="\n\n провести подбор \n\n", command =lambda: self.OpenCalculator()).grid(row=0, column=1, rowspan = 4)
        tkinter.Button(frame, text="Окно администратора", command =lambda: self.OpenAdminWindow()).grid(row=8, column=1, columnspan=1)

    def OpenAdminWindow(self):
        p = adminWindow(self.db)


    def __init__(self):
        root = tkinter.Tk()
        window1_frame = ttk.Frame(root, padding=0)
        window1_frame.grid()
        self.fillContent(window1_frame)
        root.mainloop()

if __name__ == '__main__':
    m = mainWindow()
    pass
